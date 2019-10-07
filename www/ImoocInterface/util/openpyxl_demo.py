#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author   : Li ZiHao
# @Time     : 2019/10/3 1:41
# @File     : openpyxl_demo.py

from openpyxl import load_workbook

# 1、读取excel文件
# 默认可读写，若有需要可以指定write_only和read_only为True
# wb = load_workbook('xlrd2_test.xlsx',read_only=True)
wb = load_workbook('openpyxl_test.xlsx')

# 2、读取工作表--sheet
# 获得所有sheet的名称
print(wb.get_sheet_names())
# 根据sheet名称获得sheet
a_sheet = wb.get_sheet_by_name('Sheet1')
# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet，也可以用wb.get_active_sheet()
sheet = wb.active

print(sheet.cell(row=4, column=2).coordinate)

