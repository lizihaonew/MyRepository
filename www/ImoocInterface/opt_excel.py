#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author   : Li ZiHao
# @Time     : 2019/9/20 13:03
# @File     : opt_excel.py

import datetime
import time
from openpyxl import Workbook

wb = Workbook()     # 创建文件对象
ws = wb.active      # 获取第一个sheet
ws['A1'] = 42       # A1位置写入数字
ws['B1'] = '你好' + 'hello excel'     # B1位置写入中文（unicode中文也可）
ws.append([1, 2, 3])        # 写入多个单元格
ws['A3'] = datetime.datetime.now()      # 写入一个当前时间
ws['A4'] = time.strftime('%Y-%m-%d %H:%M:%S')       # 写入一个自定义的时间

wb.save('./sample.xlsx')


